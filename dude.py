from tkinter import *
from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook('trackk.xlsx')
ws = wb['Sheet1']

wb2 = load_workbook('foodd.xlsx')
ws2 = wb2['Sheet1']

for cell in ws['A']:
    if isinstance(cell.value, datetime) and cell.value.date() == datetime.today().date() :
        today_cell = ws[cell.coordinate]

proteint_cell = today_cell.offset(column=1)
fat_cell = today_cell.offset(column=2)
carb_cell = today_cell.offset(column=3)
calories_cell = today_cell.offset(column=4)


def Check_via_macro():
    cal_macro = (int(protein_input.get()) * 4) + (int(fat_input.get()) * 9) + (int(carb_input.get()) * 4)
    text_4.configure(text= str(cal_macro) + " calories")

def Add_via_macro():
    cal_macro = (int(protein_input.get()) * 4) + (int(fat_input.get()) * 9) + (int(carb_input.get()) * 4)
    proteint_cell.value += int(protein_input.get())
    fat_cell.value += int(fat_input.get())
    carb_cell.value += int(carb_input.get())
    calories_cell.value += cal_macro
    wb.save('trackk.xlsx')
    text_4.configure(text=str(cal_macro) + " calories are added")

def Check_via_name():
    amount = int(amount_input.get())
    for cell in ws2['A']:
        if cell.value == str(name_input.get()):
            text_7.configure(text= str((cell.offset(column=4).value) * amount) + " calories")

def Add_via_name():
    amount = int(amount_input.get())
    for food in ws2['A']:
        if food.value == str(name_input.get()):
            proteint_cell.value += int(food.offset(column=1).value) * amount
            fat_cell.value += int(food.offset(column=2).value) * amount
            carb_cell.value += int(food.offset(column=3).value) * amount
            calories_cell.value += int(food.offset(column=4).value) * amount
            wb.save('trackk.xlsx')
            wb2.save('foodd.xlxs')
            text_7.configure(text= str((food.offset(column=4).value) * amount) + " calories are added")

window = Tk()
window.title("Calories Track")
window.geometry('650x350')

Title_1 = Label(window, text="Track via macronutrient", font=('TH Sarabun New',20))
Title_1.grid(row=0, column=0)

text_1 = Label(window, text="Protein: ", font=('TH Sarabun New',16))
text_1.grid(row=1, column=0, padx=5)
protein_input = Entry(window, font=('TH Sarabun New',16), width=16)
protein_input.grid(row=1, column=1)

text_2 = Label(window, text="Fat: ", font=('TH Sarabun New',16))
text_2.grid(row=2, column=0, padx=5)
fat_input = Entry(window, font=('TH Sarabun New',16), width=16)
fat_input.grid(row=2, column=1)

text_3 = Label(window, text="Carb: ", font=('TH Sarabun New',16))
text_3.grid(row=3, column=0, padx=5)
carb_input = Entry(window, font=('TH Sarabun New',16), width=16)
carb_input.grid(row=3, column=1)

text_4 = Label(window, text="", font=('TH Sarabun New',16),width=20)
text_4.grid(row=4, column=1)

check_button = Button(window, text="Check", width=8,command=Check_via_macro)
check_button.grid(row=3, column=2, padx=10)

Add_button = Button(window, text="Add", width=8,command=Add_via_macro)
Add_button.grid(row=3, column=3)

Title_2 = Label(window, text="Track via name", font=('TH Sarabun New',20))
Title_2.grid(row=5, column=0)

text_5 = Label(window, text="Food: ", font=('TH Sarabun New',16))
text_5.grid(row=6, column=0, padx=5)
name_input = Entry(window, font=('TH Sarabun New',16), width=16)
name_input.grid(row=6, column=1)

text_6 = Label(window, text="Amount: ", font=('TH Sarabun New',16))
text_6.grid(row=7, column=0, padx=5)
amount_input = Entry(window, font=('TH Sarabun New',16), width=16)
amount_input.grid(row=7, column=1)

text_7 = Label(window, text="", font=('TH Sarabun New',16),width=30)
text_7.grid(row=8, column=1)

check_button_2 = Button(window, text="Check", width=8,command=Check_via_name)
check_button_2.grid(row=7, column=2, padx=10)

Add_button_2 = Button(window, text="Add", width=8,command=Add_via_name)
Add_button_2.grid(row=7, column=3)

window.mainloop()